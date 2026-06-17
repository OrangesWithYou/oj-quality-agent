from __future__ import annotations

# 核心能力：数据模型、工具函数、Excel/JSON 读取。
# API 配置常量已独立到 config.py，如需修改接口地址/密钥/模型，请编辑那个文件。
import json
import os
import re
import time
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# 从独立配置文件导入所有 API 常量
from .config import (
    DEFAULT_OUTPUT_DIR,
    DEFAULT_POLL_INTERVAL_SEC,
    DEFAULT_POLL_TIMEOUT_SEC,
    DEFAULT_SAMPLES_PER_PROBLEM,
    JUDGE0_BASE_URL,
    JUDGE0_LANGUAGE_ID,
    PPIO_API_KEY,
    PPIO_BASE_URL,
    PPIO_MODEL,
)

# 清理代理，避免污染请求。
for _key in ["HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY", "http_proxy", "https_proxy", "all_proxy"]:
    os.environ.pop(_key, None)


# ========= 数据模型 =========

@dataclass
class SamplingConfig:
    """单次采样参数。"""
    temperature: float = 0.7
    top_p: float = 0.95
    frequency_penalty: float = 0.0
    max_tokens: int = 32768
    top_k: Optional[int] = None


@dataclass
class ProblemCase:
    """题目输入结构（应用层）。"""
    problem_id: str
    problem_text: str
    stdin: str = ""
    expected_output: Optional[str] = None
    sample_plan: List[SamplingConfig] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class RuntimeConfig:
    """运行时配置（接口参数与执行策略）。"""
    ppio_api_key: str
    ppio_base_url: str
    ppio_model: str
    judge0_base_url: str
    judge0_language_id: int
    samples_per_problem: int
    poll_interval_sec: float
    poll_timeout_sec: int
    output_dir: Path
    final_pick_strategy: str = "best_pass"


# ========= Session =========

def build_requests_session() -> requests.Session:
    """构建带重试策略的 Session。"""
    session = requests.Session()
    session.trust_env = False
    retry = Retry(
        total=3,
        connect=3,
        read=3,
        backoff_factor=1.5,
        # 429 从这里移除：由应用层（math_quiz._call_model）自行做指数退避重试，
        # 避免 urllib3 在 Session 层面抛出 RetryError 导致应用层无法捕获 429 响应。
        status_forcelist=[500, 502, 503, 504],
        allowed_methods=["GET", "POST"],
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


SESSION = build_requests_session()


# ========= 工具函数 =========

def write_json(path: Path, data: Any) -> None:
    """写入 JSON 文件（自动创建父目录）。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    """写入文本文件（自动创建父目录）。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def normalize_header(value: Any) -> str:
    """标准化表头文本：去空白、去分隔符、转小写。"""
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\s_\-:：]+", "", text)
    return text


def safe_float(value: Any, default: float) -> float:
    """安全转 float，失败时回退默认值。"""
    if value in (None, ""):
        return default
    try:
        return float(value)
    except Exception:
        return default


def safe_int(value: Any, default: Optional[int]) -> Optional[int]:
    """安全转 int，失败时回退默认值。"""
    if value in (None, ""):
        return default
    try:
        return int(float(value))
    except Exception:
        return default


def non_empty_text(value: Any) -> str:
    """将任意值转为去首尾空白的字符串。"""
    if value is None:
        return ""
    return str(value).strip()


def extract_code(text: str) -> str:
    """从模型返回文本中提取代码块；若无代码块则返回原文。"""
    text = text.strip()
    match = re.search(r"```(?:python|py)?\s*(.*?)```", text, re.S | re.I)
    return match.group(1).strip() if match else text


def find_first_diff(expected: str, actual: str) -> Optional[Dict[str, Any]]:
    """定位 expected 与 actual 的首个差异位置。"""
    expected_lines = expected.splitlines()
    actual_lines = actual.splitlines()
    max_len = max(len(expected_lines), len(actual_lines))
    for idx in range(max_len):
        exp = expected_lines[idx] if idx < len(expected_lines) else "<EOF>"
        act = actual_lines[idx] if idx < len(actual_lines) else "<EOF>"
        if exp != act:
            return {"line": idx + 1, "expected": exp, "actual": act}
    if expected != actual:
        return {"line": max_len + 1, "expected": "<TRAILING_DIFF>", "actual": "<TRAILING_DIFF>"}
    return None


def collect_failed_tests(
    judge_result: Dict[str, Any],
    stdin: str,
    expected_output: Optional[str],
) -> List[Dict[str, Any]]:
    """提取判题失败信息。"""
    failed_tests: List[Dict[str, Any]] = []
    status = judge_result.get("status") or {}
    status_id = safe_int(status.get("id"), -1)
    stdout = judge_result.get("stdout") or ""
    if expected_output is not None and status_id != 3:
        failed_tests.append({
            "type": "single_case",
            "stdin": stdin,
            "expected_output": expected_output,
            "actual_stdout": stdout,
            "first_diff": find_first_diff(expected_output, stdout),
        })
    for key in ["failed_test", "failed_tests", "test_results", "wrong_answer"]:
        if key in judge_result and judge_result[key]:
            failed_tests.append({"type": "judge_extra", "key": key, "value": judge_result[key]})
    return failed_tests


def expand_sample_plan(
    sample_plan: List[SamplingConfig],
    default_sampling: SamplingConfig,
    samples_per_problem: int,
) -> List[SamplingConfig]:
    """将样本计划补齐到固定采样次数。"""
    result: List[SamplingConfig] = []
    for cfg in sample_plan[:samples_per_problem]:
        result.append(SamplingConfig(**asdict(cfg)))
    while len(result) < samples_per_problem:
        result.append(SamplingConfig(**asdict(default_sampling)))
    return result


def pick_final_sample(samples: List[Dict[str, Any]], strategy: str) -> Dict[str, Any]:
    """按策略选择最终样本：best_pass 优先首个通过，否则取最后一个。"""
    if not samples:
        return {}
    if strategy == "best_pass":
        for sample in samples:
            if sample.get("judge_passed"):
                return sample
    return samples[-1]


def build_judge_status_summary(problem_results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """将日志信息做结构化汇总。"""
    counter: Dict[Tuple[int, str], int] = defaultdict(int)
    total_samples = 0
    passed_samples = 0
    for problem in problem_results:
        for sample in problem.get("sample_statuses", []):
            total_samples += 1
            if sample.get("judge_passed"):
                passed_samples += 1
            status_id = safe_int(sample.get("status_id"), -1)
            status_description = non_empty_text(sample.get("status_description")) or "UNKNOWN"
            counter[(int(status_id if status_id is not None else -1), status_description)] += 1
    distribution = [
        {"status_id": sid, "status_description": sdesc, "count": count}
        for (sid, sdesc), count in counter.items()
    ]
    distribution.sort(key=lambda item: (-item["count"], item["status_id"], item["status_description"]))
    failed_samples = total_samples - passed_samples
    pass_rate = (passed_samples / float(total_samples)) if total_samples else 0.0
    return {
        "total_samples": total_samples,
        "passed_samples": passed_samples,
        "failed_samples": failed_samples,
        "sample_pass_rate": pass_rate,
        "status_distribution": distribution,
    }


# ========= Excel 读取 =========

def detect_excel_columns(header_row: List[Any]) -> Dict[str, int]:
    """识别 Excel 题库字段列。"""
    aliases: Dict[str, List[str]] = {
        "problem_text": ["题目", "problem", "question", "题干"],
        "stdin": ["sampleinput", "stdin", "输入", "input"],
        "expected_output": ["sampleoutput", "expectedoutput", "输出", "output"],
        "temperature": ["temperature", "温度"],
        "top_p": ["topp"],
        "top_k": ["topk"],
        "frequency_penalty": ["frequencypenalty", "freqpenalty"],
        "max_tokens": ["maxtokens"],
    }
    columns: Dict[str, int] = {}
    normalized_headers = [normalize_header(x) for x in header_row]
    for idx, header in enumerate(normalized_headers):
        if not header:
            continue
        for field_name, candidates in aliases.items():
            if field_name in columns:
                continue
            if any(candidate in header for candidate in candidates):
                columns[field_name] = idx
    if "problem_text" not in columns:
        raise ValueError("Excel 中未找到题目列（例如：题目 / problem）")
    return columns


def read_sampling_from_row(
    row_values: List[Any],
    columns: Dict[str, int],
    default_sampling: SamplingConfig,
) -> Optional[SamplingConfig]:
    """读取单行中的采样参数（若为空行则返回 None）。"""
    def read(name: str) -> Any:
        idx = columns.get(name)
        if idx is None or idx >= len(row_values):
            return None
        return row_values[idx]

    raw_temperature = read("temperature")
    raw_top_p = read("top_p")
    raw_top_k = read("top_k")
    raw_frequency_penalty = read("frequency_penalty")
    raw_max_tokens = read("max_tokens")

    has_any = any(v not in (None, "") for v in [raw_temperature, raw_top_p, raw_top_k, raw_frequency_penalty, raw_max_tokens])
    if not has_any:
        return None

    return SamplingConfig(
        temperature=safe_float(raw_temperature, default_sampling.temperature),
        top_p=safe_float(raw_top_p, default_sampling.top_p),
        frequency_penalty=safe_float(raw_frequency_penalty, default_sampling.frequency_penalty),
        max_tokens=safe_int(raw_max_tokens, default_sampling.max_tokens) or default_sampling.max_tokens,
        top_k=safe_int(raw_top_k, default_sampling.top_k) if raw_top_k not in (None, "") else default_sampling.top_k,
    )


def load_problems_from_excel(
    excel_path: Path,
    default_sampling: SamplingConfig,
    samples_per_problem: int,
    limit: int = 0,
) -> List[ProblemCase]:
    """从 Excel 读取题目，兼容"题目起始行 + 参数续行"的格式。"""
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("读取 Excel 需要 openpyxl，请先安装：pip install openpyxl") from exc

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    columns = detect_excel_columns(header_row)

    problems: List[ProblemCase] = []
    current: Optional[ProblemCase] = None

    def finalize_current() -> None:
        nonlocal current
        if current is None:
            return
        current.sample_plan = expand_sample_plan(
            sample_plan=current.sample_plan,
            default_sampling=default_sampling,
            samples_per_problem=samples_per_problem,
        )
        problems.append(current)
        current = None

    for row_idx in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        problem_col = columns["problem_text"]
        raw_problem_text = row_values[problem_col] if problem_col < len(row_values) else None
        problem_text = non_empty_text(raw_problem_text)

        def read_cell(name: str) -> Any:
            idx = columns.get(name)
            if idx is None or idx >= len(row_values):
                return None
            return row_values[idx]

        if problem_text:
            finalize_current()
            current = ProblemCase(
                problem_id=f"problem_{len(problems) + 1:03d}",
                problem_text=problem_text,
                stdin=non_empty_text(read_cell("stdin")),
                expected_output=non_empty_text(read_cell("expected_output")) or None,
                sample_plan=[],
                metadata={"source": "excel", "row_start": row_idx},
            )
            cfg = read_sampling_from_row(row_values, columns, default_sampling)
            if cfg is not None:
                current.sample_plan.append(cfg)
        elif current is not None:
            cfg = read_sampling_from_row(row_values, columns, default_sampling)
            if cfg is not None:
                current.sample_plan.append(cfg)

    finalize_current()

    for idx, item in enumerate(problems):
        start = item.metadata.get("row_start", 2)
        if idx + 1 < len(problems):
            end = problems[idx + 1].metadata.get("row_start", ws.max_row + 1) - 1
        else:
            end = ws.max_row
        item.metadata["row_range"] = f"{start}-{end}"

    if limit > 0:
        return problems[:limit]
    return problems
